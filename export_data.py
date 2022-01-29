# -*- coding: utf-8 -*-
"""
"""
import pandas as pd
import xlsxwriter
from openpyxl import load_workbook
import os
import glob
import win32com.client as win32


def delete_file(path):
    for f in glob.iglob(path+'Output.xlsx', recursive=True):
        os.remove(f)


def create_file(file_name):
    workbook = xlsxwriter.Workbook(file_name)
    workbook.close()


def export_data(data, file_name, sheet_name):
    data = round(data, 2)
    workbook = load_workbook(file_name)

    if sheet_name in workbook.sheetnames:
        sheet = workbook.get_sheet_by_name(sheet_name)
        workbook.remove_sheet(sheet)
        workbook.save(file_name)

    writer = pd.ExcelWriter(file_name, engine='openpyxl')
    writer.book = workbook
    data.to_excel(writer, sheet_name)

    writer.save()
    writer.close()
    workbook.close()


def adjust_columns(file_name):
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    workbook = excel.Workbooks.Open(r'C:\Users\Alexandra\Documents\Python Scripts\Finist_MVP\OOP\Output.xlsx')

    for worksheet in workbook.Worksheets:
        worksheet.Columns.AutoFit()

    workbook.Save()
    excel.Application.Quit()


def delete_sheet(file_name, sheet_name):
    workbook = load_workbook(file_name)

    if sheet_name in workbook.sheetnames:
        sheet = workbook.get_sheet_by_name(sheet_name)
        workbook.remove_sheet(sheet)
        workbook.save(file_name)

    workbook.close()


def clear_products_output(file_name):
    workbook = load_workbook(file_name)
    specific_sheet = ['B_module_output', 'Лист1']
    for sheet_name in workbook.sheetnames:
        if(sheet_name not in specific_sheet):
            sheet = workbook.get_sheet_by_name(sheet_name)
            workbook.remove_sheet(sheet)
            workbook.save(file_name)

    workbook.close()
